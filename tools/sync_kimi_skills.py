# Sync local Kimi vch-* skills to the VCH v0.7.0 catalog (Phase 6, ADR-016/017/018).
# Generates/updates all 43 vch-<skill> SKILL.md files from core/VCH_HarnessCore.xlsx,
# creates the two skills missing since v6.14.0 (STATUS, CAPABILITY-DISCOVERY), and
# regenerates vch-core/references from the v0.7.0 workbook sheets.
# Run from the repository root:  python tools/sync_kimi_skills.py
import os
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
SKILLS = Path(os.path.expanduser(
    "~/Library/Application Support/kimi-desktop/daimon-share/daimon/skills"))
NEW_V = "v0.7.0"

wb = openpyxl.load_workbook(REPO / "core" / "VCH_HarnessCore.xlsx")
ws = wb["00_Skills"]
hdr = [c.value for c in ws[2]]
ix = {h: n for n, h in enumerate(hdr) if h}

skills = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row[0]:
        continue
    skills.append({h: (row[n] if row[n] is not None else "") for h, n in ix.items()})

assert len(skills) == 43, len(skills)

FOOTER = """## Shared VCH rules

Sovereign rules (SCRUB, FROZEN ORACLE, EVIDENCE, NO CAPABILITY INHERITANCE, CHARACTER POLICY),
trigger/chaining grammar, state scopes, evidence classes, persistence routing and the NEXT
footer contract are defined once in the `vch-core` skill (references/harness-rules.md).
They outrank anything in this file — read them whenever a decision depends on evidence class,
persistence or routing precedence.
"""

created, updated = [], []
for s in skills:
    sid = s["Skill_ID"]
    slug = "vch-" + sid.lower().replace("_", "-")
    d = SKILLS / slug
    d.mkdir(parents=True, exist_ok=True)
    atoms = [t.strip() for t in str(s["Trigger"]).split(" / ") if t.strip()]
    trig = ", ".join(f"'{a}'" for a in atoms)
    desc = (f"{s['Description']} Use when: {s['Use_When']} "
            f"Do NOT use when: {s['Do_Not_Use_When']} Trigger words: {trig}. "
            f"Part of the Verifiable Copilot Harness (VCH) {NEW_V} skill set; "
            f"explicit routing: [skill: {sid}].")
    chain = str(s["May_Chain_To"]).strip() or "NONE"
    chain_line = (f"May chain to: {chain}." if chain != "NONE"
                  else "May chain to: NONE (terminal).")
    body = f"""---
name: {slug}
description: "{desc}"
---

# {s['Name']} ({sid})

VCH harness skill — category: {s['Category']}, precedence: {s['Precedence']}, version: {NEW_V}.

## Behavior

{s['Key_Behavior']}

{s['Deep_Behavior']}

## Host notes

- Chat/host runtime: {s['Chat_Notes']}
- Workbook anchor: {s['Excel_Notes']}

## Chaining

{chain_line} Never chain automatically beyond the declared targets.

{FOOTER}"""
    path = d / "SKILL.md"
    (updated if path.exists() else created).append(slug)
    path.write_text(body, encoding="utf-8")

print(f"skills written: {len(created)} created {sorted(created)}")
print(f"                {len(updated)} updated")

# ---------------- vch-core/references ----------------
core = SKILLS / "vch-core" / "references"
core.mkdir(parents=True, exist_ok=True)

def sheet_rows(name):
    return [[("" if v is None else str(v)) for v in row]
            for row in wb[name].iter_rows(values_only=True)
            if any(v is not None and str(v) != "" for v in row)]

# project-rules.md
rows = sheet_rows("Project.Rules")
out = [f"# Project.Rules ({NEW_V})", ""]
for a, b, *rest in rows[2:]:
    out += [f"## {a}", b, ""]
(core / "project-rules.md").write_text("\n".join(out), encoding="utf-8")

# test-oracle.md
rows = sheet_rows("__TEST_ORACLE")
out = [f"# FROZEN TEST ORACLE ({NEW_V})", "",
       "Acceptance criteria are immutable during a run (FROZEN ORACLE sovereign rule).", "",
       "| Test_ID | Oracle_Version | Required_Field | Expected_Value | Critical | Mismatch_Result |",
       "|---|---|---|---|---|---|"]
for r in rows[2:]:
    out.append("| " + " | ".join([r[0], r[1], r[2], r[3], r[6], r[7]]) + " |")
(core / "test-oracle.md").write_text("\n".join(out) + "\n", encoding="utf-8")

# canonical-lists.md
ws_l = wb["Lists"]
headers = [ws_l.cell(2, c).value for c in range(1, ws_l.max_column + 1)]
out = [f"# Canonical Lists ({NEW_V})", "",
       "Closed enum vocabularies. Never invent new tokens; always report Field=Value.", ""]
for c, h in enumerate(headers, start=1):
    if not h:
        continue
    vals = [str(ws_l.cell(r, c).value) for r in range(3, ws_l.max_row + 1)
            if ws_l.cell(r, c).value is not None]
    out += [f"## {h}", ", ".join(vals), ""]
(core / "canonical-lists.md").write_text("\n".join(out), encoding="utf-8")

# state-schema.md
rows = sheet_rows("__STATE")
out = [f"# __STATE schema ({NEW_V})", "",
       "Canonical state fields with the immutable kernel default values. "
       "Template_Version was renamed to Schema_Version in v0.7.0 (ADR-017); "
       "pre-0.7.0 artifacts report the old name.", "",
       "| Field | Kernel default |", "|---|---|"]
for a, b, *rest in rows[2:]:
    out.append(f"| {a} | {b if b else '(empty)'} |")
(core / "state-schema.md").write_text("\n".join(out) + "\n", encoding="utf-8")

# landing-schema.md
rows = sheet_rows("00_Landing")
out = [f"# 00_Landing schema ({NEW_V})", "",
       "Landing identity fields; 00_Landing only mirrors __STATE identity "
       "(Project_ID, Project_Safe_Name) - on divergence __STATE wins.", "",
       "| Field | Kernel value |", "|---|---|"]
for a, b, *rest in rows[2:]:
    out.append(f"| {a} | {b if b else '(empty)'} |")
(core / "landing-schema.md").write_text("\n".join(out) + "\n", encoding="utf-8")

# adr.md
rows = sheet_rows("__ADR")
out = [f"# Architecture Decision Records ({NEW_V})", "",
       "| ADR_ID | Date | Status | Decision |", "|---|---|---|---|"]
for r in rows[2:]:
    out.append("| " + " | ".join([r[0], r[1], r[2], r[3]]) + " |")
(core / "adr.md").write_text("\n".join(out) + "\n", encoding="utf-8")

# glossary.md
rows = sheet_rows("__GLOSSARY")
out = [f"# Glossary ({NEW_V})", ""]
for r in rows[2:]:
    out += [f"## {r[0]}", r[1], ""]
(core / "glossary.md").write_text("\n".join(out), encoding="utf-8")

# delivery-schema.md (new in v0.7.0)
rows = sheet_rows("__DELIVERY_SCHEMA")
out = [f"# Delivery Schema ({NEW_V})", "",
       "Canonical definitions of the nine project delivery sheets (ADR-017). "
       "PROJECT-FORK generates these into revision 001; the kernel stays immutable. "
       "Out of the bootstrap read-set.", "",
       "| Sheet | Title cell | Header row |", "|---|---|---|"]
for r in rows[2:]:
    out.append("| " + " | ".join([r[0], r[1], r[2]]) + " |")
(core / "delivery-schema.md").write_text("\n".join(out) + "\n", encoding="utf-8")

# harness-rules.md: the v0.7.0 starter contract, verbatim sections
starter = (REPO / "copilot" / "copilotstart.txt").read_text(encoding="utf-8")
out = [f"# VCH Harness Rules (v0.7.0)", "",
       "Workbook-free equivalent of copilotstart.txt v0.7.0. Read on bootstrap and "
       "whenever a gate, persistence, evidence or NEXT-footer decision is made.", "",
       starter.strip(), ""]
(core / "harness-rules.md").write_text("\n".join(out), encoding="utf-8")

print("references written:", sorted(p.name for p in core.glob("*.md")))
