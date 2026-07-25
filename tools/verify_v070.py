# Read-back verification of the v0.7.0 kernel. Mirrors the harness's own lint
# philosophy: independent read-back, no self-report. Exits nonzero on any FAIL.
# Run from the repository root:  python tools/verify_v070.py
import re
import sys
import openpyxl

PATH = "core/VCH_HarnessCore.xlsx"
wb = openpyxl.load_workbook(PATH)
failures = []

def check(name, ok, detail=""):
    print(f"{'PASS' if ok else 'FAIL'}  {name}" + (f"  ({detail})" if detail and not ok else ""))
    if not ok:
        failures.append(name)

skills = wb["00_Skills"]
skill_ids = [skills.cell(r, 1).value for r in range(3, 46)]
check("SKILLCOUNT 43 skills", len([s for s in skill_ids if s]) == 43)
check("all skill versions v0.7.0",
      all(skills.cell(r, 10).value == "v0.7.0" for r in range(3, 46)))

MODES = {"HARNESS", "PROJECT", "MIGRATION"}
bad = []
for r in range(3, 46):
    v = skills.cell(r, 15).value
    if v == "ALL":
        continue
    parts = str(v).split(", ") if v else []
    if not parts or any(p not in MODES for p in parts) or ", " not in str(v) and len(parts) > 1:
        bad.append((skills.cell(r, 1).value, v))
    if re.search(r",\S|\S,", str(v)) and ", " not in str(v):
        bad.append((skills.cell(r, 1).value, v))
check("Allowed_Modes valid vs {HARNESS, PROJECT, MIGRATION}", not bad, str(bad))

row = {skills.cell(r, 1).value: r for r in range(3, 46)}
check("PROJECT-FORK modes", skills.cell(row["PROJECT-FORK"], 15).value == "HARNESS, MIGRATION")
check("PROJECT-GUIDE modes", skills.cell(row["PROJECT-GUIDE"], 15).value == "PROJECT, MIGRATION")

# chain references resolvable
ids = set(skill_ids)
bad = [cid for r in range(3, 46) for cid in str(skills.cell(r, 14).value or "").split(", ")
       if cid and cid != "NONE" and cid not in ids]
check("May_Chain_To refs resolvable", not bad, str(bad))

# trigger duplicates + prefix collisions
atoms = []
for r in range(3, 46):
    for t in str(skills.cell(r, 3).value or "").split(" / "):
        t = t.strip().lower()
        if t:
            atoms.append(t)
dupes = {a for a in atoms if atoms.count(a) > 1}
prefix = [(a, b) for a in atoms for b in atoms if a != b and b.startswith(a)]
check("trigger duplicates 0", not dupes, str(dupes))
check("trigger prefix collisions 0", not prefix, str(prefix[:5]))

# probe cell
dn = wb.defined_names.get("PROBE_CELL")
check("PROBE_CELL resolves to __STATE!$B$31", dn is not None and dn.attr_text == "__STATE!$B$31",
      str(dn and dn.attr_text))
check("__STATE!A31 == __WRITE_PROBE", wb["__STATE"]["A31"].value == "__WRITE_PROBE")

st = wb["__STATE"]
fields = {st.cell(r, 1).value: r for r in range(1, st.max_row + 1)}
check("Schema_Version present, Template_Version gone",
      "Schema_Version" in fields and "Template_Version" not in fields)
check("Schema_Version == v0.7.0", st.cell(fields["Schema_Version"], 2).value == "v0.7.0")
check("Artifact_Status == IMMUTABLE", st.cell(fields["Artifact_Status"], 2).value == "IMMUTABLE")
check("Harness version == v0.7.0", st.cell(fields["Harness version"], 2).value == "v0.7.0")

# routing oracle
ro = wb["__ROUTING_ORACLE"]
hdr_row = next(r for r in range(1, 6) if ro.cell(r, 1).value == "Test_ID")
fixtures = [ro.cell(r, 1).value for r in range(hdr_row + 1, ro.max_row + 1) if ro.cell(r, 1).value]
check("routing oracle 55 fixtures", len(fixtures) == 55, str(len(fixtures)))
check("routing Test_IDs unique", len(set(fixtures)) == len(fixtures))
bad = []
for r in range(hdr_row + 1, ro.max_row + 1):
    exp = ro.cell(r, 6).value
    if exp and exp != "NONE" and exp not in ids:
        bad.append((ro.cell(r, 1).value, exp))
check("routing Expected_Skill_ID resolvable", not bad, str(bad))
check("RT-031 rewritten", "harness kernel" in str(ro.cell(hdr_row + fixtures.index("RT-031") + 1, 4).value))
check("AV-009 rewritten", "Schema_Version" in str(ro.cell(hdr_row + fixtures.index("AV-009") + 1, 4).value))

# test oracle
to = wb["__TEST_ORACLE"]
reqs = [to.cell(r, 3).value for r in range(3, to.max_row + 1)]
check("oracle Delivery_Schema_Present", "Delivery_Schema_Present" in reqs)
ver_cells = [to.cell(r, 4).value for r in range(3, to.max_row + 1)
             if to.cell(r, 3).value in ("Harness_Version_Agreement", "Skill_Version_Format")]
check("oracle version expectations v0.7.0", all(v == "v0.7.0" for v in ver_cells), str(ver_cells))

# delivery schema
ds = wb["__DELIVERY_SCHEMA"]
names = [ds.cell(r, 1).value for r in range(3, ds.max_row + 1) if ds.cell(r, 1).value]
check("delivery schema 9 sheets", names == ["01_Plan", "02_Tickets", "03_Tasks", "04_Risks",
      "05_Script", "06_API_Reference", "07_Deployment", "08_SessionSummary", "09_Worklog"],
      str(names))
check("delivery schema headers non-empty",
      all(ds.cell(r, 2).value and ds.cell(r, 3).value for r in range(3, 12)))

# lists
li = wb["Lists"]
artifact_enum = {li.cell(r, 4).value for r in range(3, li.max_row + 1) if li.cell(r, 4).value}
filerole_enum = {li.cell(r, 16).value for r in range(3, li.max_row + 1) if li.cell(r, 16).value}
check("Artifact_Status enum: IMMUTABLE in, TEMPLATE out",
      "IMMUTABLE" in artifact_enum and "TEMPLATE" not in artifact_enum)
check("File_Role enum: no TEMPLATE", "TEMPLATE" not in filerole_enum)

# glossary
gl = wb["__GLOSSARY"]
terms = [gl.cell(r, 1).value for r in range(3, gl.max_row + 1) if gl.cell(r, 1).value]
check("glossary: Project creation mode removed", "Project creation mode" not in terms)
check("glossary: Delivery schema added", "Delivery schema" in terms)

# global scans (history sheets excluded where legitimate)
banned = re.compile(r"PROJECT.CREATION|TEMPLATE_MODE|VCH_ProjectTemplate|Template_Version")
# intentional: Project.Rules!B16 keeps the old field name once so MIGRATION mode can
# recognize pre-0.7.0 artifacts (same status as historical __ADR wording).
ALLOWED_LEFTOVERS = {"Project.Rules!B16"}
hits = []
for ws in wb.worksheets:
    if ws.title == "__ADR":  # historical records keep their original wording
        continue
    for row_ in ws.iter_rows():
        for c in row_:
            if isinstance(c.value, str) and banned.search(c.value) \
                    and f"{ws.title}!{c.coordinate}" not in ALLOWED_LEFTOVERS:
                hits.append(f"{ws.title}!{c.coordinate}: {c.value[:80]}")
check("no template/mode leftovers outside __ADR", not hits, str(hits))

v_old = []
for ws in wb.worksheets:
    if ws.title == "__ADR":
        continue
    for row_ in ws.iter_rows():
        for c in row_:
            if isinstance(c.value, str) and "v6.15.1" in c.value:
                ok_spot = (ws.title == "__STATE" and c.coordinate in ("B18", "B40")) or \
                          (ws.title == "_README" and c.coordinate == "B6")
                if not ok_spot:
                    v_old.append(f"{ws.title}!{c.coordinate}")
check("v6.15.1 only in intended historical cells", not v_old, str(v_old))

non_ascii = []
for ws in wb.worksheets:
    for row_ in ws.iter_rows():
        for c in row_:
            if isinstance(c.value, str) and any(not (32 <= ord(ch) < 127) for ch in c.value):
                non_ascii.append(f"{ws.title}!{c.coordinate}")
check("all cells printable ASCII", not non_ascii, str(non_ascii[:10]))

expected_sheets = {"_README", "00_Skills", "Project.Rules", "__STATE", "00_Landing",
                   "__TEST_ORACLE", "Lists", "__ADR", "__GLOSSARY", "__ROUTING_ORACLE",
                   "__DELIVERY_SCHEMA"}
check("sheet inventory", set(wb.sheetnames) == expected_sheets, str(wb.sheetnames))

print()
if failures:
    print(f"{len(failures)} FAILURES: {failures}")
    sys.exit(1)
print("ALL CHECKS PASS")
