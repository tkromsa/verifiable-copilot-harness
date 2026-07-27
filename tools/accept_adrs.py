# Promotion script for VCH v0.7.0 (runbook step 4). Run ONLY after all gates pass:
# structural gate, behavioral routing gate and UAT (docs/RUNBOOK_v0.7.0_gates.md).
#
# What it does:
#   1. Flips ADR-016 / ADR-017 / ADR-018 Status from Proposed to Accepted in __ADR.
#   2. Regenerates VCH_release_manifest.json (workbook hash changes with the flip).
#   3. Re-runs the read-back verification (tools/verify_v070.py).
#
# Usage (from the repository root, Windows or macOS):
#   python tools/accept_adrs.py
#   python tools/accept_adrs.py --repo C:\path\to\verifiable-copilot-harness
import argparse
import hashlib
import json
import subprocess
import sys
from pathlib import Path

import openpyxl

ADRS = ("ADR-016", "ADR-017", "ADR-018", "ADR-019")
MANIFEST_FILES = [
    "core/VCH_HarnessCore.xlsx",
    "copilot/copilotstart.txt",
    "copilot/copilot_custom_instructions.txt",
    "tools/harness_lint.ps1",
    "docs/VCH_Cheatsheet_EN.txt",
    "docs/RELEASING.md",
    "README.md",
    "LICENSE",
]

parser = argparse.ArgumentParser()
parser.add_argument("--repo", default=".", help="repository root (default: current dir)")
args = parser.parse_args()
repo = Path(args.repo).resolve()
wb_path = repo / "core" / "VCH_HarnessCore.xlsx"

# --- 1. flip ADR status -------------------------------------------------------
wb = openpyxl.load_workbook(wb_path)
ws = wb["__ADR"]
flipped, wrong = [], []
for r in range(1, ws.max_row + 1):
    adr_id = ws.cell(r, 1).value
    if adr_id in ADRS:
        status = ws.cell(r, 3).value
        if status == "Proposed":
            ws.cell(r, 3).value = "Accepted"
            flipped.append(adr_id)
        else:
            wrong.append(f"{adr_id} status={status!r}")
missing = [a for a in ADRS if a not in flipped and not any(a in w for w in wrong)]
if missing:
    print(f"FAIL: ADR rows not found: {missing}")
    sys.exit(1)
if wrong:
    print(f"NOTE: not flipped (already not Proposed): {wrong}")
if not flipped:
    print("Nothing to do: all three ADRs are already Accepted.")
else:
    wb.save(wb_path)
    print(f"flipped: {', '.join(flipped)} -> Accepted")

# --- 2. regenerate manifest ---------------------------------------------------
manifest = {}
for f in MANIFEST_FILES:
    data = (repo / f).read_bytes()
    manifest[f] = {"bytes": len(data), "sha256": hashlib.sha256(data).hexdigest()}
with open(repo / "VCH_release_manifest.json", "w") as fh:
    json.dump(manifest, fh, indent=4)
    fh.write("\n")
print(f"manifest regenerated ({len(manifest)} files)")

# --- 3. read-back verification -------------------------------------------------
result = subprocess.run(
    [sys.executable, str(repo / "tools" / "verify_v070.py")],
    cwd=repo, capture_output=True, text=True)
tail = result.stdout.strip().splitlines()[-1] if result.stdout.strip() else "(no output)"
print(f"verify_v070.py: exit={result.returncode}  {tail}")
if result.returncode != 0:
    print(result.stdout)
    print("FAIL: verification did not pass - do NOT commit or promote the release.")
    sys.exit(1)

print()
print("ALL PROMOTION STEPS PASS. Remaining manual steps:")
print('  git add -A; git commit -m "v0.7.1: gates passed, ADR-016/017/018/019 accepted"')
print("  git push origin main")
print("  gh release edit v0.7.1 --prerelease=false")
print("  (update the release notes gates section with your PASS evidence first)")
