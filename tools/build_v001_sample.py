# Simulate PROJECT-FORK: build a sample v001 project workbook from the v0.7.0
# kernel, generating the 9 delivery sheets from __DELIVERY_SCHEMA (ADR-017).
# Used by the FORK_GATE in Phase 7: the output must pass the v001-vs-schema
# check in harness_lint.ps1. Run from the repository root:
#   python tools/build_v001_sample.py [ProjectID] [ProjectName]
import shutil
import sys
from pathlib import Path

import openpyxl

KERNEL = "core/VCH_HarnessCore.xlsx"

project_id = sys.argv[1] if len(sys.argv) > 1 else "DEMO"
project_name = sys.argv[2] if len(sys.argv) > 2 else "SampleProject"
safe_name = "".join(c if c.isalnum() else "_" for c in project_name)
out = Path(f"samples/{project_id}_{safe_name}_v001.xlsx")
out.parent.mkdir(exist_ok=True)

shutil.copy(KERNEL, out)
wb = openpyxl.load_workbook(out)

# generate delivery sheets from __DELIVERY_SCHEMA
schema = wb["__DELIVERY_SCHEMA"]
for row in schema.iter_rows(min_row=3, values_only=True):
    if not row[0]:
        continue
    name, title, header, notes = row[0], row[1], row[2], row[3] if len(row) > 3 else None
    ws = wb.create_sheet(name)
    ws["A1"] = title
    for col, cell in enumerate(str(header).split(", "), start=1):
        ws.cell(2, col).value = cell
    if notes and "Seed body row:" in str(notes):
        ws["A3"] = str(notes).split("Seed body row:", 1)[1].strip()

# project identity in __STATE
st = wb["__STATE"]
fields = {st.cell(r, 1).value: r for r in range(1, st.max_row + 1)}
def set_state(field, value):
    st.cell(fields[field], 2).value = value

set_state("Project", f"{project_id} {project_name}")
set_state("File_Role", "PROJECT_WORKBOOK")
set_state("Project_ID", project_id)
set_state("Project_Safe_Name", safe_name)
set_state("Revision", "001")
set_state("Revision_Type", "PHYSICAL_ARTIFACT")
set_state("Parent_Artifact", "VCH_HarnessCore.xlsx")
set_state("Current_Artifact", out.name)
set_state("Active_Artifact", out.name)
set_state("Physical_Artifact", out.name)
set_state("Artifact_Status", "CURRENT")
set_state("Created_From", "VCH_HarnessCore.xlsx v0.7.0")

# mirror identity on 00_Landing (landing mirrors __STATE)
land = wb["00_Landing"]
labels = {land.cell(r, 1).value: r for r in range(1, land.max_row + 1)}
for label, value in (("Project ID", project_id), ("Project Name", project_name)):
    if label in labels:
        land.cell(labels[label], 2).value = value

wb.save(out)
print(f"created {out}")

# immediate read-back: v001 delivery sheets must match __DELIVERY_SCHEMA exactly
wb2 = openpyxl.load_workbook(out)
schema2 = wb2["__DELIVERY_SCHEMA"]
errors = []
for row in schema2.iter_rows(min_row=3, values_only=True):
    if not row[0]:
        continue
    name, title, header = row[0], row[1], row[2]
    if name not in wb2.sheetnames:
        errors.append(f"missing sheet {name}")
        continue
    ws = wb2[name]
    if ws["A1"].value != title:
        errors.append(f"{name}: title {ws['A1'].value!r} != {title!r}")
    actual = ", ".join(str(c.value) for c in ws[2] if c.value is not None)
    if actual != header:
        errors.append(f"{name}: header {actual!r} != {header!r}")
if errors:
    print("FORK_GATE FAIL:")
    for e in errors:
        print(" -", e)
    sys.exit(1)
print("FORK_GATE PASS: v001 delivery sheets match __DELIVERY_SCHEMA")
