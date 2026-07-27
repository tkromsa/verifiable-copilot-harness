# Build VCH_HarnessCore.xlsx v0.7.1 from the v0.7.0 kernel (ADR-019):
# embed copilot/copilotstart.txt as the _STARTER sheet, bump versions, add the
# STARTER oracle row, glossary term and ADR-019. Run from the repository root:
#   python tools/build_v071.py
import openpyxl

PATH = "core/VCH_HarnessCore.xlsx"
STARTER_TXT = "copilot/copilotstart.txt"
OLD_V = "v0.7.0"
NEW_V = "v0.7.1"

wb = openpyxl.load_workbook(PATH)
changes = []

# --- 1. global version bump (exact + substring), excluding historical spots --
EXCLUDE_CELLS = {("_README", "B6")}  # provenance keeps original release identity
for ws in wb.worksheets:
    if ws.title == "__ADR":
        continue  # historical records keep original wording
    for row in ws.iter_rows():
        for c in row:
            if not isinstance(c.value, str) or OLD_V not in c.value:
                continue
            if (ws.title, c.coordinate) in EXCLUDE_CELLS:
                continue
            c.value = c.value.replace(OLD_V, NEW_V)
            changes.append(f"{ws.title}!{c.coordinate}: version bump")

# --- 2. lineage cells --------------------------------------------------------
st = wb["__STATE"]
fields = {st.cell(r, 1).value: r for r in range(1, st.max_row + 1)}
st.cell(fields["Created_From"], 2).value = f"VCH_HarnessCore.xlsx {OLD_V}"
st.cell(fields["Source_Refs"], 2).value = f"VCH_HarnessCore.xlsx {OLD_V}; ADR-019"
changes.append("__STATE!B18/B40: lineage -> v0.7.0 + ADR-019")

# --- 3. _STARTER sheet (generated from the txt source of truth) ---------------
text = open(STARTER_TXT, encoding="utf-8").read()
lines = text.split("\n")
while lines and lines[-1] == "":
    lines.pop()
assert not any(any(ord(ch) < 32 or ord(ch) > 126 for ch in ln) for ln in lines), \
    "starter contains non-printable or non-ASCII characters"
if "_STARTER" in wb.sheetnames:
    del wb["_STARTER"]
ws = wb.create_sheet("_STARTER", 1)  # right after _README
for i, ln in enumerate(lines, start=1):
    ws.cell(i, 1).value = ln
changes.append(f"_STARTER: created from {STARTER_TXT} ({len(lines)} rows)")

# --- 4. _README read-set -------------------------------------------------------
rd = wb["_README"]
rd["B13"] = ("_STARTER, Project.Rules, 00_Landing, __STATE, 00_Skills, __TEST_ORACLE, "
             "__ADR, __GLOSSARY. __ROUTING_ORACLE and __DELIVERY_SCHEMA are NOT in the "
             "bootstrap read-set; PROJECT-FORK reads __DELIVERY_SCHEMA at fork.")
changes.append("_README!B13: read-set starts with _STARTER")

# --- 5. oracle row -------------------------------------------------------------
to = wb["__TEST_ORACLE"]
to.append(["STARTER", "1.0", "Starter_Sheet_Present", "YES", "NONE",
           "DETERMINISTIC_SCAN", "YES", "FAIL",
           "_STARTER sheet exists and its rows reproduce copilotstart.txt exactly; "
           "harness_lint compares the sheet against the repo file."])
changes.append("__TEST_ORACLE: appended STARTER / Starter_Sheet_Present")

# --- 6. glossary term ----------------------------------------------------------
gl = wb["__GLOSSARY"]
gl.append(["Starter sheet",
           "Sheet _STARTER embedded in the kernel carrying the copilotstart contract; "
           "generated from copilot/copilotstart.txt at build time and read first on "
           "load vch. The repo txt remains the source of truth.",
           "_STARTER", "separate starter attachment", "ADR-019", "2026-07-27"])
changes.append("__GLOSSARY: appended term 'Starter sheet'")

# --- 7. ADR-019 ----------------------------------------------------------------
adr = wb["__ADR"]
adr.append(["ADR-019", "2026-07-27", "Proposed",
            "Embed copilotstart.txt as the _STARTER sheet in the kernel; the txt stays "
            "the repo source of truth and the lint asserts sheet equals file",
            "After ADR-017 the starter was the only behavior text outside the artifact. "
            "Version drift between the attached starter and the kernel was unverifiable "
            "at runtime, and two attachments per session were the last avoidable "
            "onboarding step.",
            "Keep a separate starter txt; embed a generated _STARTER sheet; embed and "
            "delete the txt from the repo.",
            "A version-locked artifact removes starter/kernel drift by construction, the "
            "same argument as ADR-012 for the routing oracle. Keeping the txt as the "
            "build source preserves git-diffable review; the sheet is generated, never "
            "hand-edited.",
            "Attach flow becomes one workbook per session; custom instructions start the "
            "bootstrap read-set with _STARTER; harness_lint gains the STARTER "
            "sheet-versus-file check; __TEST_ORACLE gains Starter_Sheet_Present; "
            "copilot_custom_instructions.txt stays outside because it configures Copilot "
            "itself and cannot be attached. All components bump to v0.7.1.",
            "Delete the _STARTER sheet and attach copilot/copilotstart.txt again; the "
            "txt is kept in the repo unchanged in role.",
            "Owner request 2026-07-27; tools/build_v071.py."])
changes.append("__ADR: appended ADR-019 (Proposed)")

wb.save(PATH)
print(f"saved {PATH}; {len(changes)} changes")
for c in changes[:6]:
    print(" -", c)
print(f" ... ({len(changes)} total)")
